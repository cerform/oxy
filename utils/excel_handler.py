import os
from openpyxl import Workbook, load_workbook

# Создаём папку 'data', если её нет
os.makedirs("data", exist_ok=True)

# Путь к файлу Excel
file_path = "data/usage_data.xlsx"

# Заголовки (из названий полей формы)
HEADERS = ["Full Name", "Worker ID", "Department", "Usage Date", "Amount Used"]


def save_to_excel(data):
    """Функция для сохранения данных в Excel"""

    # Если файл не существует — создаём новый с заголовками
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)  # Записываем заголовки
        wb.save(file_path)

    # Открываем существующий файл и добавляем данные
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append(data)  # Добавляем строку с данными
    wb.save(file_path)


# 🛠 Пример использования:
if __name__ == "__main__":
    new_data = ["John Doe", "12345", "Lab A", "2025-03-03", "15L"]
    save_to_excel(new_data)
    print("Данные успешно сохранены!")
